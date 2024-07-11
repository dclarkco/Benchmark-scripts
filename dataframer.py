##################################
# Benchmark Mineral Intelligence #
# ---------DCLARK 2024---------- #
#         dataframer.py          #
##################################

# This script contains functions to clean and process data, and to extract information from text
# It also contains functions to retrieve ISO country codes and currencies

# Cleaned of private keys and references, this code was useful for a single project, and is now redundant following the finalization of my contract's work.
# The fruits of this code are reflected in the Benchmark Pro mineral database, and it's utility is no longer required.
# It has been cleared as an example of the work I did during my time at BMI.
# This is the original "War-Chest" box of scripts, as described by my last senior developer.

import pandas as pd, numpy as np, regex as re
import os

debug = True
export = True

def data_framer(file_path, sheetname = None, clean = True, header_row = 6):
	# Check if the file exists
	if not os.path.isfile(file_path):
		raise FileNotFoundError(f"The file '{file_path}' does not exist.")

	# Get the file extension (e.g., ".csv" or ".xlsx")
	file_extension = os.path.splitext(file_path)[-1].lower()

	# Check if it's a CSV file
	if file_extension == '.csv':
		try:
			# Read the CSV file into a DataFrame
			df = pd.read_csv(file_path)
		except pd.errors.EmptyDataError:
			raise ValueError(f"The CSV file '{file_path}' is empty.")
		except pd.errors.ParserError:
			raise ValueError(f"Unable to parse the CSV file '{file_path}'.")
	
	# Check if it's an Excel file
	elif file_extension in ['.xls', '.xlsx', '.xlsm']:
		try:
			# Read the Excel file into a DataFrame
			df = pd.read_excel(file_path, sheet_name = sheetname, header = header_row)
			return df


		except pd.errors.EmptyDataError:
			raise ValueError(f"The Excel file '{file_path}' is empty.")
		except pd.errors.XLRDError:
			raise ValueError(f"Unable to parse the Excel file '{file_path}'.")
	
	# If it's neither CSV nor Excel, raise an error
	else:
		raise ValueError(f"The file '{file_path}' is not a valid CSV or Excel file.")
	df.replace('nan', np.nan, inplace=True)
	if clean:
		df = data_cleaner(df)
		#print(df.columns)
	return df

def data_cleaner(df):
	debug = True

	if debug:
		print("Null before clean:  {}%".format(round(df.isnull().sum().sum()/len(df), 2)))
		print("Duplicate before clean:  {}%".format(round(df.duplicated().sum().sum()/len(df), 2)))

	df.replace(["", " ", None, pd.NA,"NaN"], np.nan, inplace=True) # replace na values with numpy nan

	threshold_empty_columns = 0.9  # Adjust this threshold as needed
	threshold_empty_rows = 0.9  # Adjust this threshold as needed
	# Drop columns that are mostly empty
	df = df.dropna(axis = 1, thresh = (df.shape[1]*threshold_empty_columns))
	# Drop rows that are mostly empty
	df = df.dropna(axis = 0, thresh = (df.shape[0]*threshold_empty_rows))
	#df = df.loc[:, ~df.columns.str.match('\d{4}', na=False)].copy() # drop columns with dates
	#df = df.loc[:, ~df.columns.duplicated()].copy()         # drop duplicate columns
	#df = df.drop_duplicates()                               # drop duplicate values
	# Drop rows with NaN values in the first N (8) columns
	df = df.dropna(subset=df.columns[:8], how='any')
	df.columns = df.columns.astype(str)
	df.reset_index(drop=True, inplace=True)

	if debug:
		print("Null after clean: {}%".format(round(df.isnull().sum().sum()/len(df), 2)))
		print("Duplicate after clean: {}%".format(round(df.duplicated().sum().sum()/len(df), 2)))
		print(df.tail())
		print(list(df.columns))

	return df

def detect_headers(df, detect_word): # looks for a specific word, and sets the headers to the row after that word
	for index, row in df.iterrows():
		clean_row_values = list(map(clean_and_lower, row.values))
		if detect_word in clean_row_values:
			df.columns = row.values
			df = df.iloc[index+1:]

			break  # Stop iterating once "Status" is found
	return df

	# for index, row in df.iterrows(): # alternate method which tried to detect the first row with less than 10% NAN values. didnt work.
	# 	empty = row.isna().mean()
	# 	print("EMPTY % at row {}: {}".format(index, empty))
	# 	if empty < 0.1: # if row is less than 10% NAN, set as first row of dataframe (headers)
	# 		return index
	# 		print('---------------------------------', headerLn)

def clean_xlsx(path): # removes all formatting from an excel file
	import openpyxl

	# Load the Excel file
	file_path = path
	workbook = openpyxl.load_workbook(file_path)

	# Iterate through all sheets in the workbook
	for sheet_name in workbook.sheetnames:
		sheet = workbook[sheet_name]

		# Clear all formatting for each cell in the sheet
		for row in sheet.iter_rows():
			for cell in row:
				cell.font = openpyxl.styles.Font()
				cell.fill = openpyxl.styles.PatternFill()
				cell.border = openpyxl.styles.Border()
				cell.alignment = openpyxl.styles.Alignment()

	# Save the modified workbook
	workbook.save(f"{file_path}_cleaned.xlsx")

def clean_and_lower(s): #strip and lower text, useful for pandas .apply methods
	s = str(s)
	return s.strip().lower() 

def extract_first_substring(s, delimiters):
	# Split the string at the specified delimiters, and return the first non-empty part
	return next((part for part in re.split(delimiters, s) if part), None)

def remove_punctuation(s): # removes punctuation from a string, useful for pandas .apply methods
	punctuation_pattern = r'[^\w\s]'
	# Use the re.sub() function to replace punctuation with an empty string
	cleaned_string = re.sub(punctuation_pattern, '', str(s))
	
	return cleaned_string

def clean_text(text, custom_stopwords = None): # removes punctuation and stopwords from a string - used for cleaning company corpus of junk text

	if text is None:
		return ''

	from nltk.corpus import stopwords
	from nltk.tokenize import word_tokenize
	import re, nltk

	#nltk.download('punkt')
	#nltk.download("stopwords")

	# Remove punctuation and non-alphanumeric characters
	text = re.sub(r'[^a-zA-Z\s]', '', text)
	# Tokenize the text and remove stop words
	stop_words = set(stopwords.words("english"))

	if custom_stopwords is not None:
		stop_words.update(custom_stopwords)

	word_tokens = word_tokenize(text.lower())
	filtered_words = [word for word in word_tokens if word not in stop_words]
	# Join the filtered words back into a single string
	cleaned_text = " ".join(filtered_words).title()
	
	return cleaned_text

def get_ISO_countries(): # gets a list of ISO3166 countries and their codes
	from isocodes import countries

	countryDF = pd.DataFrame.from_records(countries.items)
	for country in countries.items:
		print(country['name'],'|' ,country['alpha_3'])

	if export:
		countryDF.to_csv("reference/ISO3166_Countries.csv", encoding='utf-8')

	return countryDF

def get_ISO_subdivisions(): # gets a list of ISO3166 subdivisions and their codes
	from isocodes import subdivisions_countries as subdivisions

	subdivisionDF = pd.DataFrame(columns = ['subdivision_name', 'subdivision_code', 'subdivision_type', 'country_code', 'country_name'])
	for subdivision in subdivisions.items:
		#print(subdivision, '|', subdivision['name'],'|' ,subdivision['code'])
		alpha_2_code = subdivision["code"].split('-')[0]
		alpha_3_code = get_country_code(alpha_2_code, input_mode='alpha_2', output_mode='alpha_3')
		country_name = get_country_code(alpha_2_code, input_mode='alpha_2', output_mode='name')
		subdivisionDF.loc[len(subdivisionDF.index)] = ({'subdivision_name': subdivision['name'], 'subdivision_code': subdivision['code'], 'subdivision_type': subdivision['type'], 'country_code': alpha_3_code, 'country_name': country_name})

		print(get_country_code(alpha_2_code, input_mode='alpha_2', output_mode='name'), subdivision)
		
	if export:
		subdivisionDF.to_csv("reference/ISO3166_Subdivisions.csv", encoding='utf-8')

	return subdivisionDF

def get_ISO_currencies(): # gets a list of ISO4217 currencies and their codes
	from isocodes import currencies

	currencyDF = pd.DataFrame.from_records(currencies.items)
	for currency in currencies.items:
		print(currency,'|' ,currency)

	if export:
		currencyDF.to_csv("reference/ISO4217_Countries.csv", encoding='utf-8')

	return currencyDF

def convert_to_ISO(file, sheet = 'Li Supply'): # converts a series of countries in a worksheet column to ISO3166 codes

	df = pd.read_excel(file, sheet_name = sheet)

	# rename mis-named countries in the database, so they match the scheme and are correctly matched to ISO
	replacement_dict = {'USA': 'United States', 'UK': 'United Kingdom', 'DRC': 'Democratic Republic of the Congo', 'Czech Republic ': 'Czechia', 'UAE': 'United Arab Emirates'}
	df['Country'] = df['Country'].replace(replacement_dict)

	df['Country Code'] = df['Country'].apply(get_country_code)

	df.to_csv('{}_ISO3166_remediated.csv'.format(sheet), index=False)

def get_countries(): # gets a list of countries and their regions from restcountries.com (alternate source, ISO is preferred.) -> Useful for getting extra metadata about countries
	import requests

	# Fetch data from the REST Countries API
	url = "https://restcountries.com/v3.1/all"
	response = requests.get(url)

	# Check if the request was successful (status code 200)
	if response.status_code == 200:
		# Load JSON data into a Python object
		countries_data = response.json()

		# Extract relevant information and create a list of dictionaries
		countries_list = []
		for country_data in countries_data:
			if country_data["name"]["common"] == "China":
				batt_demand_region = "China"
			elif country_data["continents"][0] == "North America":
				batt_demand_region = "North America"
			elif country_data["continents"][0] == "South America":
				batt_demand_region = "South America"
			elif country_data["continents"][0] == "Oceania":
				batt_demand_region = "APAC (excl. China)"
			elif country_data["region"] == "Asia":
				batt_demand_region = "APAC (excl. China)"
			elif country_data["region"] == "Oceania":
				batt_demand_region = "APAC (excl. China)"
			elif country_data["region"] == "Europe":
				batt_demand_region = "Europe"
			elif country_data["region"] == "Africa":
				batt_demand_region = "Other"
			else:
				batt_demand_region = "Other"

			country_info = {
				"name": country_data["name"]["common"],
				"official_name": country_data["name"]["official"],
				"alpha_2": country_data["cca2"],
				"alpha_3": country_data["cca3"],
				"capital": ", ".join(country_data.get("capital", "")),
				"continents": ", ".join(country_data.get("continents", "")),
				"region": country_data.get("region", ""),
				"batt_demand_region": batt_demand_region,
				"subregion": country_data.get("subregion", ""),
				"population": country_data.get("population", ""),
				"timezones": ", ".join(country_data.get("timezones", "")),
				"currency": ", ".join(country_data.get("currencies", "")),
				"latitude": country_data.get("latlng", [])[0],
				"longitude": country_data.get("latlng", [])[1],
			}
			countries_list.append(country_info)

		# Convert the list of dictionaries to a Pandas DataFrame
		countries_df = pd.DataFrame(countries_list)
		countries_df.sort_values(["batt_demand_region", "subregion", "name"], ascending=[True, True, True], inplace=True)
 
		# Display the DataFrame
		print(countries_df)

		print(f"Regions: {countries_df['region'].unique()} \nContinents: {countries_df['continents'].unique()} \nSubregions: {countries_df['subregion'].unique()}")
		countries_df.to_csv('reference/countries_regions.csv', index=False)

	else:
		print(f"Failed to retrieve data. Status code: {response.status_code}")

# Finds the best matching ISO country code for a given country name or alpha_2 abbreviation (can be used in tandem with convert_to_ISO to remediate country names in a dataset)
def get_country_code(country_name, input_mode = 'name', output_mode = 'alpha_3'):
	# input/output modes: name, alpha_2, alpha_3
	from thefuzz import fuzz

	iso_df = pd.read_csv('reference/ISO3166_Countries.csv')
	max_ratio = 0
	best_match = None
	
	for iso_index, iso_row in iso_df.iterrows():

		ratio = fuzz.token_sort_ratio(country_name, iso_row[input_mode])

		if ratio > max_ratio:
			max_ratio = ratio
			best_match = iso_row[output_mode]

	return best_match

def split_and_expand(input_string):
	if pd.notna(input_string):
		# Split by "(" and "/", and "," and ")" and expand the result into a list
		parts = [part.strip() for part in re.split(r'[\/\(\)]', input_string) if part.strip()]
		# Update the existing value with the first substring
		primary_name = parts[0] if parts else ""
		# Create a 'Notes' column with the remaining substrings
		notes = ', '.join(parts[1:]) if len(parts) > 1 else ''        
		return primary_name, notes
	else:
		return "", ""

def generate_slug(input_string):
	# Remove all non-alphanumeric characters
	slug = re.sub(r'[^a-zA-Z0-9\s]', '', input_string)
	# Replace all whitespace with a hyphen
	slug = re.sub(r'\s', '-', slug)
	# Convert the slug to lowercase
	slug = slug.lower()
	return slug